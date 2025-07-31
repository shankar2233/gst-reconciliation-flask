import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelHandler:
    def __init__(self):
        self.upload_folder = 'uploads'
    
    def create_sample_format(self):
        """Create sample Excel format"""
        # Sample data for Tally sheet
        tally_data = {
            'GSTIN of supplier': ['27AABCU9603R1ZX', '27AABCU9603R1ZY', '27AABCU9603R1ZZ'],
            'Supplier': ['ABC Private Ltd', 'XYZ Industries', 'GHI Enterprises'],
            'Invoice number': ['INV-0001', 'INV-0002', 'INV-0003'],
            'Invoice Date': ['15-04-2024', '20-04-2024', '25-04-2024'],
            'Invoice Value': [118000, 177000, 112000],
            'Rate': [18, 18, 12],
            'Taxable Value': [100000, 150000, 100000],
            'Integrated Tax': [0, 27000, 0],
            'Central Tax': [9000, 0, 6000],
            'State/UT tax': [9000, 0, 6000],
            'Cess': [0, 0, 0]
        }
        
        # Sample data for GSTR-2A sheet
        gstr_data = {
            'GSTIN of supplier': ['27AABCU9603R1ZX', '27AABCU9603R1ZZ', '27AABCU9603R1ZW'],
            'Supplier': ['ABC Private Limited', 'DEF Corporation', 'GHI Enterprises'],
            'Invoice number': ['INV-0001', 'INV-0004', 'INV-0003'],
            'Invoice Date': ['15-04-2024', '25-04-2024', '25-04-2024'],
            'Invoice Value': [118000, 89600, 112000],
            'Rate': [18, 12, 12],
            'Taxable Value': [100000, 80000, 100000],
            'Integrated Tax': [0, 0, 0],
            'Central Tax': [9000, 4800, 6000],
            'State/UT tax': [9000, 4800, 6000],
            'Cess': [0, 0, 0]
        }
        
        df_tally = pd.DataFrame(tally_data)
        df_gstr = pd.DataFrame(gstr_data)
        
        # Create Excel file
        sample_path = os.path.join(self.upload_folder, 'sample_format.xlsx')
        
        with pd.ExcelWriter(sample_path, engine='openpyxl') as writer:
            df_tally.to_excel(writer, sheet_name='Tally', index=False)
            df_gstr.to_excel(writer, sheet_name='GSTR-2A', index=False)
        
        return sample_path
    
    def generate_report(self, results, session_id):
        """Generate comprehensive Excel report"""
        report_path = os.path.join(self.upload_folder, f'{session_id}_report.xlsx')
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary data
        summary_data = [
            ['GST Reconciliation Report'],
            [''],
            ['Metric', 'Value'],
            ['Total Tally Records', results['summary']['total_tally_records']],
            ['Total GSTR Records', results['summary']['total_gstr_records']],
            ['Matched Records', results['summary']['matched_records']],
            ['Discrepancies', results['summary']['discrepancies']],
            ['Unmatched Tally', results['summary']['unmatched_tally']],
            ['Unmatched GSTR', results['summary']['unmatched_gstr']],
            ['Match Percentage', f"{results['summary']['match_percentage']}%"]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Style the summary sheet
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['B3'].font = Font(bold=True)
        
        # Matched Records sheet
        if results['matched_records']:
            ws_matched = wb.create_sheet(title="Matched Records")
            matched_df = pd.DataFrame(results['matched_records'])
            
            for r in dataframe_to_rows(matched_df, index=False, header=True):
                ws_matched.append(r)
        
        # Discrepancies sheet
        if results['discrepancies']:
            ws_discrepancies = wb.create_sheet(title="Discrepancies")
            discrepancies_df = pd.DataFrame(results['discrepancies'])
            
            for r in dataframe_to_rows(discrepancies_df, index=False, header=True):
                ws_discrepancies.append(r)
        
        # Unmatched Tally sheet
        if results['unmatched_tally']:
            ws_unmatched_tally = wb.create_sheet(title="Unmatched Tally")
            unmatched_tally_df = pd.DataFrame(results['unmatched_tally'])
            
            for r in dataframe_to_rows(unmatched_tally_df, index=False, header=True):
                ws_unmatched_tally.append(r)
        
        # Unmatched GSTR sheet
        if results['unmatched_gstr']:
            ws_unmatched_gstr = wb.create_sheet(title="Unmatched GSTR")
            unmatched_gstr_df = pd.DataFrame(results['unmatched_gstr'])
            
            for r in dataframe_to_rows(unmatched_gstr_df, index=False, header=True):
                ws_unmatched_gstr.append(r)
        
        wb.save(report_path)
        return report_path
